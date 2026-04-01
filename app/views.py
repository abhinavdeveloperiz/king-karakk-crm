from django.shortcuts import render,redirect,get_object_or_404
from .forms import BranchCreateForm,BranchEditForm,AdminBranchTransactionCreateForm, BranchTransactionCreateForm
from .models import Branch,User,Transaction, Branch
from django.contrib.auth import update_session_auth_hash


from django.contrib.auth.hashers import make_password
from django.contrib.auth import authenticate, login,logout
from django.contrib.auth.decorators import login_required

from django.contrib import messages

from django.db.models import Sum, Case, When, DecimalField
from django.db.models.functions import TruncDate
from decimal import Decimal

from django.core.paginator import Paginator
from django.db.models import Q

from datetime import date,datetime, time
import calendar


from collections import defaultdict
from django.utils import timezone
from django.utils.timezone import now
import json








def admin_and_branch_login(request):
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect("admin_dashboard")
        elif getattr(request.user, "is_branch", False):
            return redirect("branch_dashboard")
        else:
            logout(request)

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is None:
            messages.error(request, "Invalid username or password.")
            return redirect("login")   

        if not (user.is_superuser or getattr(user, "is_branch", False)):
            messages.error(request, "You are not allowed to login here.")
            return redirect("login")

        login(request, user)

        return redirect(
            "admin_dashboard" if user.is_superuser else "branch_dashboard"
        )

    return render(request, "auth/login.html")





@login_required(login_url="login")
def change_superadmin_credentials(request):
    user = request.user

    if not user.is_superuser:
        return redirect("login")

    if request.method == "POST":
        old_password = request.POST.get("old_password")
        new_username = request.POST.get("new_username", "").strip()
        new_password = request.POST.get("new_password")
        confirm_password = request.POST.get("confirm_password")

        # 1️⃣ Old password check
        if not user.check_password(old_password):
            messages.error(request, "Old password is incorrect.")
            return redirect("change_credentials")

        # 2️⃣ Password confirmation
        if new_password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return redirect("change_credentials")

        # 3️⃣ No-change validation
        username_unchanged = new_username == user.username
        password_unchanged = user.check_password(new_password)

        if username_unchanged and password_unchanged:
            messages.error(
                request,
                "New username and password must be different from the current credentials."
            )
            return redirect("change_credentials")
        

        if new_username and not username_unchanged:
            user.username = new_username

        # 5️⃣ Update password if changed
        if not password_unchanged:
            user.set_password(new_password)

        user.save()
        update_session_auth_hash(request, user)

        messages.success(request, "Admin credentials updated successfully.")
        return redirect("admin_profile")

    return render(request, "auth/change_credentials.html")





def admin_and_branch_logout(request):
    logout(request)
    return redirect("login")



@login_required(login_url="login")
def admin_profile(request):
    current_year = now().year

    total_branches = Branch.objects.exclude(name__iexact="office").count()
    transactions = Transaction.objects.filter(created_on__year=current_year)


    # this year transaction 
    totals = transactions.aggregate(
        total_purchase=Sum('amount', filter=Q(transaction_type='PURCHASE')),
        total_expense=Sum('amount', filter=Q(transaction_type='EXPENSE')),
        total_cashbalance=Sum('amount', filter=Q(transaction_type='CASHBALANCE')),
    )

    total_purchase = totals['total_purchase'] or Decimal('0')
    total_expense = totals['total_expense'] or Decimal('0')
    total_cashbalance = totals['total_cashbalance'] or Decimal('0')

    total_sales = total_purchase + total_expense + total_cashbalance


    # today transactions
    today = timezone.localdate()
    today_transactions = Transaction.objects.filter(created_on__date=today)

    today_totals = today_transactions.aggregate(
        today_purchase=Sum('amount', filter=Q(transaction_type='PURCHASE')),
        today_expense=Sum('amount', filter=Q(transaction_type='EXPENSE')),
        today_cashbalance=Sum('amount', filter=Q(transaction_type='CASHBALANCE')),
    )

    today_purchase = today_totals['today_purchase'] or Decimal('0')
    today_expense = today_totals['today_expense'] or Decimal('0')
    today_cashbalance = today_totals['today_cashbalance'] or Decimal('0')

    today_total_sales = today_purchase + today_expense + today_cashbalance


    context = {
        'year': current_year,
        'total_purchase': total_purchase,
        'total_expense': total_expense,
        'total_cashbalance': total_cashbalance,
        'total_sales': total_sales,
        'total_branches': total_branches,
        'today_purchase': today_purchase,
        'today_expense': today_expense,
        'today_cashbalance': today_cashbalance,
        'today_total_sales': today_total_sales,
    }

    return render(request, 'owner/admin_profile.html', context)





@login_required(login_url="login")
def admin_dashboard(request):
    return render(request, 'owner/admin_dashboard.html')






@login_required(login_url="login")
def daily_sales_report(request):
    today = timezone.localdate()
    current_month = today.month
    current_year = today.year

    from calendar import monthrange
    _, last_day = monthrange(current_year, current_month)
    start_date = date(current_year, current_month, 1)
    end_date = date(current_year, current_month, last_day)

    # Get all branches
    branches = Branch.objects.exclude(name__iexact="office").order_by("name")

    # Generate daily data for the month
    daily_data = []
    for day in range(1, last_day + 1):
        day_date = date(current_year, current_month, day)
        start_of_day = timezone.make_aware(datetime.combine(day_date, time.min))
        end_of_day = timezone.make_aware(datetime.combine(day_date, time.max))

        day_branches = []
        total_day_sales = Decimal('0.00')

        for branch in branches:
            # Get transactions for this branch on this day
            transactions = Transaction.objects.filter(
                branch=branch,
                created_on__range=(start_of_day, end_of_day)
            )

            # Calculate total sales for this branch on this day (PURCHASE + EXPENSE + CASHBALANCE)
            branch_sales = transactions.filter(transaction_type__in=['PURCHASE', 'EXPENSE', 'CASHBALANCE']).aggregate(
                total=Sum('amount', default=Decimal('0.00'))
            )['total'] or Decimal('0.00')

            day_branches.append({
                'branch': branch,
                'total_sales': branch_sales,
            })

            total_day_sales += branch_sales

        daily_data.append({
            'date': day_date,
            'day': day,
            'branches': day_branches,
            'total_sales': total_day_sales,
        })

    # Calculate monthly totals
    total_month_sales = sum(day['total_sales'] for day in daily_data)

    # Calculate branch monthly totals
    branch_monthly_totals = []
    for branch in branches:
        branch_total_sales = Decimal('0.00')
        for day_data in daily_data:
            for branch_data in day_data['branches']:
                if branch_data['branch'].id == branch.id:
                    branch_total_sales += branch_data['total_sales']
        branch_monthly_totals.append({
            'branch': branch,
            'total_sales': branch_total_sales,
        })

    context = {
        "current_month": current_month,
        "current_year": current_year,
        "daily_data": daily_data,
        "branches": branches,
        "branch_monthly_totals": branch_monthly_totals,
        "month_name": date(current_year, current_month, 1).strftime('%B %Y'),
        "total_month_sales": total_month_sales,
    }

    return render(request, "owner/daily_sales_report.html", context)




@login_required(login_url="login")
def export_monthly_report(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    from io import BytesIO

    today = timezone.localdate()
    current_month = today.month
    current_year = today.year

    # Calculate start and end of the current month
    from calendar import monthrange
    _, last_day = monthrange(current_year, current_month)
    start_date = date(current_year, current_month, 1)
    end_date = date(current_year, current_month, last_day)

    # Get all branches
    branches = Branch.objects.exclude(name__iexact="office").order_by("name")

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Monthly Report - {date(current_year, current_month, 1).strftime('%B %Y')}"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Headers
    headers = ['Date'] + [branch.name for branch in branches] + ['Total']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Generate data for the month
    row_num = 2
    for day in range(1, last_day + 1):
        day_date = date(current_year, current_month, day)
        start_of_day = timezone.make_aware(datetime.combine(day_date, time.min))
        end_of_day = timezone.make_aware(datetime.combine(day_date, time.max))

        day_branches = []
        total_day_sales = Decimal('0.00')

        for branch in branches:
            # Get transactions for this branch on this day
            transactions = Transaction.objects.filter(
                branch=branch,
                created_on__range=(start_of_day, end_of_day)
            )

            # Calculate total sales for this branch on this day (PURCHASE + EXPENSE + CASHBALANCE)
            branch_sales = transactions.filter(transaction_type__in=['PURCHASE', 'EXPENSE', 'CASHBALANCE']).aggregate(
                total=Sum('amount', default=Decimal('0.00'))
            )['total'] or Decimal('0.00')

            day_branches.append({
                'branch': branch,
                'total_sales': branch_sales,
            })

            total_day_sales += branch_sales

        # Write data to Excel
        ws.cell(row=row_num, column=1, value=day_date.strftime('%m/%d/%Y'))

        col_num = 2
        for branch_data in day_branches:
            ws.cell(row=row_num, column=col_num, value=float(branch_data['total_sales']) if branch_data['total_sales'] > 0 else None).number_format = '#,##0.00'
            col_num += 1

        ws.cell(row=row_num, column=col_num, value=float(total_day_sales) if total_day_sales > 0 else None).number_format = '#,##0.00'

        # Apply borders
        total_cols = 1 + len(branches) + 1
        for col in range(1, total_cols + 1):
            ws.cell(row=row_num, column=col).border = border

        row_num += 1

    # Monthly Total Row
    ws.cell(row=row_num, column=1, value='Monthly Total')

    total_month_sales = Decimal('0.00')
    col_num = 2
    for branch in branches:
        branch_total_sales = Decimal('0.00')
        for day in range(1, last_day + 1):
            day_date = date(current_year, current_month, day)
            start_of_day = timezone.make_aware(datetime.combine(day_date, time.min))
            end_of_day = timezone.make_aware(datetime.combine(day_date, time.max))
            transactions = Transaction.objects.filter(
                branch=branch,
                created_on__range=(start_of_day, end_of_day)
            )
            branch_total_sales += transactions.filter(transaction_type__in=['PURCHASE', 'EXPENSE', 'CASHBALANCE']).aggregate(total=Sum('amount', default=Decimal('0.00')))['total'] or Decimal('0.00')
        ws.cell(row=row_num, column=col_num, value=float(branch_total_sales) if branch_total_sales > 0 else None).number_format = '#,##0.00'
        total_month_sales += branch_total_sales
        col_num += 1

    ws.cell(row=row_num, column=col_num, value=float(total_month_sales) if total_month_sales > 0 else None).number_format = '#,##0.00'

    # Apply borders and bold to total row
    total_cols = 1 + len(branches) + 1
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.border = border
        cell.font = Font(bold=True)

    # Auto-adjust column widths
    for col_num, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max(max_length + 2, 15) if col_num > 1 else (max_length + 2)  # Minimum width of 15 for data columns
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="monthly_report_{current_year}_{current_month:02d}.xlsx"'

    return response




from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from datetime import datetime, date, time
from decimal import Decimal
from django.db.models import Sum
from calendar import monthrange


@login_required(login_url="login")
def Admin_cashflow(request):

    today = timezone.localdate()

    selected_month = int(request.GET.get("month", today.month))
    selected_year = int(request.GET.get("year", today.year))
    selected_branch_id = request.GET.get("branch")

    # Month range
    _, last_day = monthrange(selected_year, selected_month)

    branches = Branch.objects.all().order_by("name")
    selected_branch = None
    if selected_branch_id:
        selected_branch = branches.filter(id=selected_branch_id).first()
    branches_to_show = [selected_branch] if selected_branch else branches

    cashflow_data = []

    for day in range(1, last_day + 1):
        day_date = date(selected_year, selected_month, day)

        start_of_day = timezone.make_aware(datetime.combine(day_date, time.min))
        end_of_day = timezone.make_aware(datetime.combine(day_date, time.max))

        day_transactions = Transaction.objects.filter(
            created_on__range=(start_of_day, end_of_day)
        )

        branch_data = []

        for branch in branches_to_show:

            branch_tx = day_transactions.filter(branch=branch)

            purchase = branch_tx.filter(transaction_type="PURCHASE").aggregate(
                total=Sum("amount")
            )["total"] or Decimal("0.00")

            expense = branch_tx.filter(transaction_type="EXPENSE").aggregate(
                total=Sum("amount")
            )["total"] or Decimal("0.00")

            cashbalance = branch_tx.filter(transaction_type="CASHBALANCE").aggregate(
                total=Sum("amount")
            )["total"] or Decimal("0.00")

            # Transfers
            transfer_out = branch_tx.filter(
                transaction_type="TRANSFER"
            ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

            transfer_in = day_transactions.filter(
                transaction_type="TRANSFER",
                target_branch=branch
            ).aggregate(total=Sum("amount"))["total"] or Decimal("0.00")

            # 💡 SALES LOGIC (as you asked)
            sales = purchase + expense + cashbalance + transfer_in - transfer_out

            branch_data.append({
                "branch": branch,
                "purchase": purchase,
                "expense": expense,
                "cashbalance": cashbalance,
                "transfer_in": transfer_in,
                "transfer_out": transfer_out,
                "sales": sales,
            })

        cashflow_data.append({
            "date": day_date,
            "branches": branch_data,
        })

    # 🔁 Separate Transfer Table Data
    transfers = Transaction.objects.filter(
        transaction_type="TRANSFER",
        created_on__month=selected_month,
        created_on__year=selected_year
    )
    if selected_branch:
        transfers = transfers.filter(Q(branch=selected_branch) | Q(target_branch=selected_branch))
    transfers = transfers.select_related("branch", "target_branch").order_by("-created_on")

    context = {
        "cashflow_data": cashflow_data,
        "branches": branches,
        "selected_branch": selected_branch,
        "transfers": transfers,
        "selected_month": selected_month,
        "selected_year": selected_year,
    }

    return render(request, "owner/admin_cashflow.html", context)


# -------------------------------------------------------------- 

@login_required(login_url="login")
def branch_list(request):
    branches = Branch.objects.select_related('user').all().order_by("name")
    return render(request, 'owner/admin_branch_list.html', {"branches": branches})



@login_required(login_url="login")
def branch_add(request):
    if request.method == "POST":
        form = BranchCreateForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Branch Added successfully")
            return redirect("branch_list") 
        else:
            messages.error(request, "Failed to Add")
    else:
        form = BranchCreateForm()

    return render(request, 'owner/admin_branch_add.html', {"form": form})



@login_required(login_url="login")
def branch_edit(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    user = branch.user

    if request.method == "POST":
        form = BranchEditForm(request.POST)
        if form.is_valid():
            user.username = form.cleaned_data["username"]

            new_password = form.cleaned_data["password"]
            if new_password.strip() != "":
                user.password = make_password(new_password)

            user.save()

            # Update branch
            branch.name = form.cleaned_data["name"]
            branch.location = form.cleaned_data["location"]
            branch.save()
            messages.success(request, f"Branch {branch.name} Updated")
            return redirect("branch_list")
        else:
            messages.error(request, "Failed to Update")

    else:
        form = BranchEditForm(initial={
            "name": branch.name,
            "location": branch.location,
            "username": user.username,
        })

    return render(request, "owner/admin_branch_edit.html", {
        "form": form,
        "branch": branch
    })



@login_required(login_url="login")
def branch_delete(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)
    user = branch.user

    if request.method == "POST":
        entered_username = request.POST.get("username")
        entered_password = request.POST.get("password")

        if entered_username == user.username and user.check_password(entered_password):
            user.delete()
            messages.success(request, "Branch deleted successfully")
            return redirect("branch_list")
        else:
            messages.error(request, "Invalid username or password")

    return render(request, "owner/branch_delete_confirm.html", {"branch": branch})








@login_required(login_url="login")
def branch_detail(request, branch_id):
    branch = get_object_or_404(Branch, id=branch_id)

    today = timezone.localdate()
    
    # Get selected date from request, default to today
    selected_date_str = request.GET.get("date")
    if selected_date_str:
        try:
            from datetime import datetime as dt
            selected_date = dt.strptime(selected_date_str, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            selected_date = today
    else:
        selected_date = today

    start_of_day = timezone.make_aware(datetime.combine(selected_date, time.min))
    end_of_day = timezone.make_aware(datetime.combine(selected_date, time.max))
   
    transactions = Transaction.objects.filter(
        branch=branch,
        created_on__gte=start_of_day, created_on__lte=end_of_day
    ).order_by("-created_on")

    sales_transactions = transactions.filter(transaction_type="SALE")
    expense_transactions = transactions.filter(transaction_type="EXPENSE")
    purchase_transactions = transactions.filter(transaction_type="PURCHASE")
    cashbalance_transactions = transactions.filter(transaction_type="CASHBALANCE")
    transfer_transactions = transactions.filter(transaction_type="TRANSFER")

    totals = transactions.aggregate(
        total_sales=Sum(
            Case(
                When(transaction_type="SALE", then="amount"),
                default=Decimal("0.00"),
                output_field=DecimalField()
            )
        ),
        total_expense=Sum(
            Case(
                When(transaction_type="EXPENSE", then="amount"),
                default=Decimal("0.00"),
                output_field=DecimalField()
            )
        ),
        total_purchase=Sum(
            Case(
                When(transaction_type="PURCHASE", then="amount"),
                default=Decimal("0.00"),
                output_field=DecimalField()
            )
        ),
        total_cashbalance=Sum(
            Case(
                When(transaction_type="CASHBALANCE", then="amount"),
                default=Decimal("0.00"),
                output_field=DecimalField()
            )
        ),
        total_transfer=Sum(
            Case(
                When(transaction_type="TRANSFER", then="amount"),
                default=Decimal("0.00"),
                output_field=DecimalField()
            )
        ),
    )

    total_sales = totals["total_sales"] or Decimal("0.00")
    total_expense = totals["total_expense"] or Decimal("0.00")
    total_purchase = totals["total_purchase"] or Decimal("0.00")
    total_cashbalance = totals["total_cashbalance"] or Decimal("0.00")
    total_transfer = totals["total_transfer"] or Decimal("0.00")

    # Calculate received transfers (SALE transactions from transfers)
    received_transfers_raw = sales_transactions.filter(description__startswith="Transfer from")
    total_received_transfer = received_transfers_raw.aggregate(
        total=Sum("amount", default=Decimal("0.00"))
    )["total"] or Decimal("0.00")

    # Process received transfers to extract source branch names
    received_transfers = []
    for transfer in received_transfers_raw:
        source_branch = "Unknown"
        if "Transfer from" in transfer.description:
            # Extract branch name from "Transfer from BranchName: description"
            desc_part = transfer.description[14:]  # Remove "Transfer from "
            if ":" in desc_part:
                source_branch = desc_part.split(":")[0].strip()
        
        received_transfers.append({
            'transaction': transfer,
            'source_branch': source_branch
        })

    # New logic: expected_sales = purchase + expense + cashbalance
    balance_total = total_purchase + total_expense + total_cashbalance

    context = {
        "branch": branch,
        "today": today,
        "selected_date": selected_date,
        "sales_transactions": sales_transactions,
        "expense_transactions": expense_transactions,
        "purchase_transactions": purchase_transactions,
        "cashbalance_transactions": cashbalance_transactions,
        "transfer_transactions": transfer_transactions,
        "received_transfers": received_transfers,
        "total_sales": total_sales,
        "total_expense": total_expense,
        "total_purchase": total_purchase,
        "total_cashbalance": total_cashbalance,
        "total_transfer": total_transfer,
        "total_received_transfer": total_received_transfer,
        "balance_total": balance_total,
    }

    return render(request, "owner/admin_branch_detail.html", context)






@login_required(login_url="login")
def admin_add_transaction_to_branch(request, branch_id):
    if not request.user.is_superuser:
        return redirect("login")

    branch = get_object_or_404(Branch, id=branch_id)

    if request.method == "POST":
        form = AdminBranchTransactionCreateForm(request.POST)
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.branch = branch
            selected_date = form.cleaned_data['created_on']
            transaction.created_on = timezone.make_aware(datetime.combine(selected_date, timezone.now().time()))
            transaction.full_clean()
            transaction.save()

            # If it's a transfer, create corresponding transaction for target branch
            if transaction.transaction_type == "TRANSFER":
                target_branch = transaction.target_branch
                # Create a SALE transaction for the target branch
                Transaction.objects.create(
                    branch=target_branch,
                    transaction_type="SALE",
                    amount=transaction.amount,
                    description=f"Transfer from {branch.name}: {transaction.description or ''}",
                    created_on=transaction.created_on
                )

            messages.success(request, f"Transaction added successfully to {branch.name}.")
            return redirect("branch_detail", branch_id=branch_id)
    else:
        form = AdminBranchTransactionCreateForm()

    # Limit target_branch choices to other branches
    form.fields['target_branch'].queryset = Branch.objects.exclude(id=branch.id)

    return render(
        request,
        "owner/admin_add_transaction.html",
        {"form": form, "branch": branch}
    )













@login_required(login_url="login")
def branch_dashboard(request):

    try:
        branch = request.user.branch_account
    except Branch.DoesNotExist:
        branch = None

    today = timezone.localdate()

    total_sales = total_expense = total_purchase = total_cashbalance = balance_total = Decimal("0.00")
    chart_data = {"labels": [], "values": []}

    if branch:
        start_of_day = timezone.make_aware(datetime.combine(today, time.min))
        end_of_day = timezone.make_aware(datetime.combine(today, time.max))
        transactions = branch.transactions.filter(created_on__gte=start_of_day, created_on__lte=end_of_day)

        totals = transactions.aggregate(
            total_sales=Sum(
                Case(
                    When(transaction_type="SALE", then="amount"),
                    default=Decimal("0.00"),
                    output_field=DecimalField()
                )
            ),
            total_expense=Sum(
                Case(
                    When(transaction_type="EXPENSE", then="amount"),
                    default=Decimal("0.00"),
                    output_field=DecimalField()
                )
            ),
            total_purchase=Sum(
                Case(
                    When(transaction_type="PURCHASE", then="amount"),
                    default=Decimal("0.00"),
                    output_field=DecimalField()
                )
            ),
            total_cashbalance=Sum(
                Case(
                    When(transaction_type="CASHBALANCE", then="amount"),
                    default=Decimal("0.00"),
                    output_field=DecimalField()
                )
            ),
        )

        total_sales = totals["total_sales"] or Decimal("0.00")
        total_expense = totals["total_expense"] or Decimal("0.00")
        total_purchase = totals["total_purchase"] or Decimal("0.00")
        total_cashbalance = totals["total_cashbalance"] or Decimal("0.00")

        # New logic: sales_of_day = purchase + expense + cashbalance
        balance_total = total_purchase + total_expense + total_cashbalance

        chart_data = {
            "labels": ["Purchase", "Expenses", "Cash Balance"],
            "values": [
                float(total_purchase),
                float(total_expense),
                float(total_cashbalance),
            ],
        }

    context = {
        "branch": branch,
        "today": today,
        "total_sales": total_sales,
        "total_expense": total_expense,
        "total_purchase": total_purchase,
        "total_cashbalance": total_cashbalance,
        "balance_total": balance_total,
        "chart_data_json": json.dumps(chart_data),
    }

    return render(request, "branch/branch_dashboard.html", context)






@login_required(login_url="login")
def branch_expense_sales_list(request):
    try:
        branch = request.user.branch_account
    except Branch.DoesNotExist:
        messages.error(request, "You do not have a branch assigned.")
        return redirect("branch_dashboard")

    today = timezone.localdate()
    
    # Get selected date from request, default to today
    selected_date_str = request.GET.get("date")
    if selected_date_str:
        try:
            from datetime import datetime as dt
            selected_date = dt.strptime(selected_date_str, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            selected_date = today
    else:
        selected_date = today

    start_of_day = timezone.make_aware(datetime.combine(selected_date, time.min))
    end_of_day = timezone.make_aware(datetime.combine(selected_date, time.max))

    transactions = branch.transactions.filter(
        created_on__gte=start_of_day, created_on__lte=end_of_day
    ).order_by("-created_on")

    sales_transactions = transactions.filter(transaction_type="SALE")
    expense_transactions = transactions.filter(transaction_type="EXPENSE")
    purchase_transactions = transactions.filter(transaction_type="PURCHASE")
    cashbalance_transactions = transactions.filter(transaction_type="CASHBALANCE")
    transfer_transactions = transactions.filter(transaction_type="TRANSFER")

    totals = transactions.aggregate(
     
        total_expense=Sum(
            Case(
                When(transaction_type="EXPENSE", then="amount"),
                default=Decimal("0.00"),
                output_field=DecimalField()
            )
        ),
        total_purchase=Sum(
            Case(
                When(transaction_type="PURCHASE", then="amount"),
                default=Decimal("0.00"),
                output_field=DecimalField()
            )
        ),
        total_cashbalance=Sum(
            Case(
                When(transaction_type="CASHBALANCE", then="amount"),
                default=Decimal("0.00"),
                output_field=DecimalField()
            )
        ),
        total_transfer=Sum(
            Case(
                When(transaction_type="TRANSFER", then="amount"),
                default=Decimal("0.00"),
                output_field=DecimalField()
            )
        ),
    )

    total_expense = totals["total_expense"] or Decimal("0.00")
    total_purchase = totals["total_purchase"] or Decimal("0.00")
    total_cashbalance = totals["total_cashbalance"] or Decimal("0.00")
    total_transfer = totals["total_transfer"] or Decimal("0.00")

    # Calculate received transfers (SALE transactions from transfers)
    received_transfers_raw = sales_transactions.filter(description__startswith="Transfer from")
    total_received_transfer = received_transfers_raw.aggregate(
        total=Sum("amount", default=Decimal("0.00"))
    )["total"] or Decimal("0.00")

    # Process received transfers to extract source branch names
    received_transfers = []
    for transfer in received_transfers_raw:
        source_branch = "Unknown"
        if "Transfer from" in transfer.description:
            # Extract branch name from "Transfer from BranchName: description"
            desc_part = transfer.description[14:]  # Remove "Transfer from "
            if ":" in desc_part:
                source_branch = desc_part.split(":")[0].strip()
        
        received_transfers.append({
            'transaction': transfer,
            'source_branch': source_branch
        })

    # New logic: sales_of_day = purchase + expense + cashbalance
    balance_total = total_purchase + total_expense + total_cashbalance

    context = {
        "branch": branch,
        "today": today,
        "selected_date": selected_date,
        "sales_transactions": sales_transactions,
        "expense_transactions": expense_transactions,
        "purchase_transactions": purchase_transactions,
        "cashbalance_transactions": cashbalance_transactions,
        "transfer_transactions": transfer_transactions,
        "received_transfers": received_transfers,
        "total_expense": total_expense,
        "total_purchase": total_purchase,
        "total_cashbalance": total_cashbalance,
        "total_transfer": total_transfer,
        "total_received_transfer": total_received_transfer,
        "balance_total": balance_total,
    }

    return render(request, "branch/branch_expense_sales.html", context)




@login_required(login_url="login")
def branch_expense_sales_entry(request):

    try:
        branch = request.user.branch_account
    except Branch.DoesNotExist:
        messages.error(request, "No branch assigned")
        return redirect("branch_dashboard")

    if request.method == "POST":
        form = BranchTransactionCreateForm(request.POST, branch=branch)

        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.branch = branch

            selected_date = form.cleaned_data["created_on"]
            transaction.created_on = timezone.make_aware(
                datetime.combine(selected_date, datetime.now().time())
            )

            transaction.save()

            if transaction.transaction_type == "TRANSFER":
                Transaction.objects.create(
                    branch=transaction.target_branch,
                    transaction_type="SALE",
                    amount=transaction.amount,
                    description=f"Transfer from {branch.name}",
                    created_on=transaction.created_on
                )

            messages.success(request, "Saved successfully")
            return redirect("branch_expense_sales_list")

        else:
            print(form.errors)  # DEBUG

    else:
        form = BranchTransactionCreateForm(branch=branch)

    return render(request, "branch/branch_expense_sales_add.html", {"form": form})




@login_required(login_url="login")
def branch_transfer(request):
    try:
        branch = request.user.branch_account
    except Branch.DoesNotExist:
        messages.error(request, "You do not have a branch assigned. Please contact admin.")
        return redirect("branch_dashboard")

    if request.method == "POST":
        # Create transfer transaction
        target_branch_id = request.POST.get("target_branch")
        amount = request.POST.get("amount")
        description = request.POST.get("description", "")
        created_on = request.POST.get("created_on")

        try:
            target_branch = Branch.objects.get(id=target_branch_id)
            amount = Decimal(amount)
            created_on = datetime.strptime(created_on, "%Y-%m-%d").date()
            created_on_aware = timezone.make_aware(datetime.combine(created_on, timezone.now().time()))

            # Create transfer out transaction
            transfer_out = Transaction.objects.create(
                branch=branch,
                transaction_type="TRANSFER",
                target_branch=target_branch,
                amount=amount,
                description=description,
                created_on=created_on_aware
            )

            # Create corresponding sale for target branch
            Transaction.objects.create(
                branch=target_branch,
                transaction_type="SALE",
                amount=amount,
                description=f"Transfer from {branch.name}: {description}",
                created_on=created_on_aware
            )

            messages.success(request, f"Successfully transferred BD {amount} to {target_branch.name}.")
            return redirect("branch_expense_sales_list")

        except (Branch.DoesNotExist, ValueError, Decimal.InvalidOperation) as e:
            messages.error(request, "Invalid data provided.")

    # Get other branches
    other_branches = Branch.objects.exclude(id=branch.id)
    today = timezone.now().date()

    return render(
        request,
        "branch/branch_transfer.html",
        {
            "branch": branch,
            "other_branches": other_branches,
            "today": today
        }
    )


@login_required(login_url="login")
def admin_transfer(request, branch_id):
    if not request.user.is_superuser:
        return redirect("login")

    branch = get_object_or_404(Branch, id=branch_id)

    if request.method == "POST":
        # Create transfer transaction
        target_branch_id = request.POST.get("target_branch")
        amount = request.POST.get("amount")
        description = request.POST.get("description", "")
        created_on = request.POST.get("created_on")

        try:
            target_branch = Branch.objects.get(id=target_branch_id)
            amount = Decimal(amount)
            created_on = datetime.strptime(created_on, "%Y-%m-%d").date()
            created_on_aware = timezone.make_aware(datetime.combine(created_on, timezone.now().time()))

            # Create transfer out transaction
            transfer_out = Transaction.objects.create(
                branch=branch,
                transaction_type="TRANSFER",
                target_branch=target_branch,
                amount=amount,
                description=description,
                created_on=created_on_aware
            )

            # Create corresponding sale for target branch
            Transaction.objects.create(
                branch=target_branch,
                transaction_type="SALE",
                amount=amount,
                description=f"Transfer from {branch.name}: {description}",
                created_on=created_on_aware
            )

            messages.success(request, f"Successfully transferred BD {amount} from {branch.name} to {target_branch.name}.")
            return redirect("branch_detail", branch_id=branch_id)

        except (Branch.DoesNotExist, ValueError, Decimal.InvalidOperation) as e:
            messages.error(request, "Invalid data provided.")

    # Get other branches
    other_branches = Branch.objects.exclude(id=branch.id)
    today = timezone.now().date()

    return render(
        request,
        "owner/admin_transfer.html",
        {
            "branch": branch,
            "other_branches": other_branches,
            "today": today
        }
    )


@login_required(login_url="login")
def branch_profile(request):
    branch = Branch.objects.only("id", "name", "location").get(user=request.user)

    recent_transactions = (
        Transaction.objects
        .filter(branch=branch)
        .only(
            "transaction_type",
            "purchase_category",
            "expense_category",
            "cashbalance_category",
            "amount",
            "created_on",
        )
        .order_by("-created_on")[:3]
    )

    return render(
        request,
        "branch/branch_profile.html",
        {
            "branch": branch,
            "recent_transactions": recent_transactions,
        },
    )
